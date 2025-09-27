import pandas as pd
from openpyxl import load_workbook
import os

def analyze_excel_file(filename):
    print(f"\n=== Analysis of {filename} ===")
    if not os.path.exists(filename):
        print(f"File {filename} not found")
        return

    try:
        wb = load_workbook(filename)
        print(f"Sheets: {wb.sheetnames}")

        for sheet in wb.sheetnames:
            try:
                df = pd.read_excel(filename, sheet_name=sheet)
                print(f"\n{sheet}:")
                print(f"  Shape: {df.shape}")
                print(f"  Columns: {list(df.columns)}")
                if not df.empty:
                    print(f"  First 3 rows:")
                    print(df.head(3).to_string())
            except Exception as e:
                print(f"  Error reading sheet {sheet}: {e}")

    except Exception as e:
        print(f"Error analyzing {filename}: {e}")

if __name__ == "__main__":
    analyze_excel_file("data_final.xlsx")
    analyze_excel_file("ModeloPLS_final.xlsx")